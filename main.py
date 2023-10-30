import pandas as pd

from scipy.optimize import minimize
import numpy as np
from matplotlib import pyplot as plt
import paneltime as pt
import statsmodels.api as sm
import pickle
import openpyxl
from openpyxl.styles import Font


def main():
    df = handle_data()

    save_transitions(df)


    X = sm.add_constant(df[['normalized_rank', 'normalized_rank sq', 'semidemdummy', 'demdummy']])
    mesh(df, 'normalized_rank')
    with open('mesh.pd','rb') as f:
        d = pickle.load(f)
    df_res = create_df(d)
    calc_transition(df_res, d)
    #res = OLS(df, X)
    pres = panel_reg(df)
    plotting(df, pres)

def draw_network():
   
def save_transitions(df):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ['ifhpol', 'normalized_rank']:
        thresholds = [
            [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9],
            [0.2,0.4,0.6,0.8],
            [1/3,2/3]]
        ws_markov = wb.create_sheet(f'markov {name}')
        ws_trans  = wb.create_sheet(f'trans {name}')
        lens = np.array([len(i)+2 for i in thresholds], dtype=int)
        markov = np.full((sum(lens+2), max(lens)+3), '', dtype='<U20')
        trans = np.full((sum(lens+2), max(lens)+3), '', dtype='<U20')
        markov[0,0] = f'Markov matrix {name} (frequencies)'
        trans[0,0] = f'Matrix of transitions {name} (number of cases)'
        sheets = []
        for i, t in enumerate(thresholds):
            mkv, tr, t = calc_group_movements_gen(df, t, name)
            row = sum((lens+2)[:i])+2
            n = len(t)
            for x,m,s,perc in [(markov, mkv, ws_markov,  True), 
                                (trans, tr,  ws_trans,   False)]:
                x[row+1,0] = 'To interval'
                x[row-1,2] = 'From interval'
                x[row,n+2] = 'sum'
                x[row,2:n+2] = t
                x[row+1:row + n+1,1] = t 
                x[row+1:row + n+1,2:n+2] = m
                x[row+1:row + n+1,n+2] = np.sum(m,axis=1)
                populate_ws(s,x)
                sheets.append(s)
                set_fmt(s, 'font', Font(bold = True),row-1, 2, row+1, len(x[0]))
                if perc:
                    fmt = "0 %"
                else:
                    fmt = "_-* # ##0_-;-* # ##0_-;_-* \"-\"??_-;_-@_-"
                set_fmt(s, 'number_format', fmt,row+1, 2, row + n+1, n+2)

        for s in sheets:
            set_fmt(s, 'font', Font(bold = True), 0, 0, len(markov), 2)

    
    #populate_ws(wb.create_sheet('data'), df)
    wb.save(f'markov.xlsx')
                   
def populate_ws(ws, x):
    """Sets a range in the worksheet from a 2D list or numpy array."""
    if type(x)==pd.DataFrame:
        x = np.vstack((x.columns,np.array(x)))
    for i, row in enumerate(x):
        for j, value in enumerate(row):
            if value == '':
                value = None
            try:
                value = float(value)
            except:
                pass
            if value == float('inf'):
                value = 'inf'
            ws.cell(i+1, j+1).value = value
    return ws

def set_fmt(ws, property, fmt, row, col, endrow = None, endcol = None):
    if endcol is None:
        endcol = col +1
    if endrow is None:
        endrow = col +1
    for i in range(row, endrow):
        for j in range(col, endcol):
            setattr(ws.cell(i+1,j+1),property,fmt)




def create_df(d):
    df = pd.DataFrame(d).T
    df.columns = ['threshold 1','threshold 2','g1', 'g2','g3', 'g2->g1','g3->g1','g1->g2','g3->g2','g2->g3','g1->g3','g1->g1','g2->g2','g3->g3']
    return df

def markov(g1, g2, g3, mg21,mg31,mg12,mg32,mg23,mg13, mg11, mg22, mg33):
    mkv = np.array([ 
        [mg11/g1, mg12/g1,  mg13/g1], 
        [mg21/g2, mg22/g2 , mg23/g2], 
        [mg31/g3, mg32/g3 , mg33/g3]
        ])
    movements = np.sum(mkv[1]**0.5)
    test = np.array([(mg11+mg12+mg13)/g1, (mg21+mg22+mg23)/g2,(mg31+mg32+mg33)/g3])#should all equal 1
    if (not np.all(test==np.array([1,1,1])) and (g1*g2*g3>0)) or (movements>6):
        a=0
    return mkv, movements

def plotting(df, res):
    df = df[['normalized_rank', 'change_pred']]
    df = df.sort_values('normalized_rank')
    df.plot('normalized_rank', 'change_pred')
    plt.show()
    x = np.linspace(0,1,100)
    a = res.ll.args.args_d
    plt.plot(x,a['beta'][1,0]+ a['beta'][2,0]*np.abs(x-0.5)**2)
    plt.title('prediction of normalized_rank and normalized_rank_sq only')
    plt.show()
    
def handle_data():
    df = pd.read_csv("datasimplified.csv")
    df = get_annual_rank(df)

    df = pd.DataFrame(df)
    df['normalized_rank sq'] = np.abs(df['normalized_rank']-0.5)**2
    df['demdummy'] = (df['normalized_rank']==1.0)*1.0
    df['semidemdummy'] = (df['normalized_rank']>0.5)*1.0
    return df
    
def panel_reg(df=None):
    pt.options.pqdkm.set([1,1,0,1,1])
    pt.options.fixed_random_time_eff.set(2)
    pt.options.fixed_random_group_eff.set(2)
    pt.options.tolerance.set(0.000001)
    
    res = pt.execute('abs_one_year_change~normalized_rank+normalized_rank sq+semidemdummy+demdummy', df, 
               ID = 'country', T = 'year'
               )
    print(res)
    df['change_pred']  = res.predict()['in-sample predicted Y']
    return res
    
    
def OLS(df, X):
    model = sm.OLS(df['abs_one_year_change'], X)
    results = model.fit()
    df['change_pred'] = results.predict()
    print(results.summary()) 
    return results

def calc_transition(df, d):
    "Selects thresholds that mazximize difference in group volatility"
    
    maxd = np.max(list(d))
    print(f"max: {maxd}, t:{d[maxd]}")
    mkv, movements = markov(*tuple(d[maxd][2:]))
    a=0
        
def get_annual_rank(df):
    df = pd.DataFrame(df)
    df = df.dropna()
    df = df.sort_values(['year', 'ifhpol'])
    # Group the data by 'year'
    grouped = df.groupby('year')
    # Add a new column 'rank' with the ranking for each identifier within each year
    df['rank'] = grouped['ifhpol'].rank(ascending=True)

    # Calculate the maximum rank for each year
    max_rank = grouped['rank'].transform('max')

    # Calculate the normalized rank for each identifier within each year
    df['normalized_rank'] = df['rank'] / max_rank
    df['max_rank'] = max_rank
    

    df = df.sort_values(['country', 'year'])

    grouped = df.groupby('country')

    # Calculate the one-year change in the 'variable' column
    df['rank_from'] = df['rank'].shift(1)
    df['ifhpol_from'] = df['ifhpol'].shift(1)
    df['totdurny2_from'] = df['totdurny2'].shift(1)
    df['abs_one_year_change'] = np.abs(grouped['normalized_rank'].diff())
    df['normalized_rank_from'] = df['rank_from'] / df['max_rank']

    # Drop the first row for each group as it will have a NaN value for the change
    df = df.dropna()
    # Print the resulting DataFrame
    return df

def calc_group_movements(df, t1, t2, name):
    g1 = df[(df[name]<=t1)][name + '_from']
    g2 = df[(df[name]>t1)&(df[name]<t2)][name + '_from']
    g3 = df[df[name]>=t2][name + '_from']


    mg12 = np.sum((g1>t1)&(g1<t2))      #g1->g2
    mg13 = np.sum((g1>t2))              #g1->g3
    mg21 = np.sum((g2<t1))              #g2->g1
    mg23 = np.sum((g2>t2))              #g3->g3
    mg32 = np.sum((g3<t2)&(g3>t1))      #g2->g2
    mg31 = np.sum((g3<t1))              #g3->g1

    mg11 = np.sum((g1<=t1))             #g1->g1
    mg22 = np.sum((g2>t1)&(g2<t2))      #g2->g2
    mg33 = np.sum((g3>=t2))             #g3->g3



    return len(g1), len(g2), len(g3), mg21,mg31,mg12,mg32,mg23,mg13, mg11, mg22, mg33

def calc_group_movements_gen(df, thresholds, name):
    gs = []
    n = len(thresholds) +1
    thresholds = np.array(thresholds)*np.max(df[name])
    thresholds = [0] + list(thresholds) + [np.inf]
    thresholds = np.column_stack((thresholds, np.roll(thresholds, -1)))[:-1]
    m = np.zeros((n,n))
    markov = np.zeros((n,n))
    for i, (t0,t1) in enumerate(thresholds):
        g = df[(df[name]>=t0)&(df[name]<t1)][name + '_from']
        gs.append(g)

        for j, (to_t0, to_t1) in enumerate(thresholds):
            m[i,j] = np.sum((g>=to_t0)&(g<to_t1))
            markov[i,j] = m[i,j]/len(g)

    return markov, m, thresholds[:,1]

def mesh(df, name):
    d = {}
    x = 0.01
    mint = np.min(df[name])
    for t1 in np.arange(mint,1.0,x):
        for t2 in np.arange(t1+x,1.0,x):
            mkv, m = calc_group_movements_gen(df, [t1, t2], name)
            movements = np.abs(mkv-(1.0/len(mkv[0])))
            if not np.any(np.isnan(mkv)):
                d[movements] =[t1, t2]

main()