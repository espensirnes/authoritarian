import pandas as pd

import numpy as np
import paneltime as pt
import statsmodels.api as sm
import pickle
import openpyxl
from openpyxl.styles import Font
import os
os.chdir(os.path.dirname(__file__))
import networkdrawing as nw

   

def save_transitions(df):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in ['ifhpol', 'normalized_rank']:
        thresholds = [
            [0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8,0.9],
            [0.2,0.4,0.6,0.8],
            [1/3,2/3]]
        ws_markov = wb.create_sheet(f'markov {name}')
        ws_trans  = wb.create_sheet(f'trans {name}')
        lens = np.array([len(i)+2 for i in thresholds], dtype=int)
        mrkov = np.full((sum(lens+2), max(lens)+3), '', dtype='<U20')
        trans = np.full((sum(lens+2), max(lens)+3), '', dtype='<U20')
        mrkov[0,0] = f'Markov matrix {name} (frequencies)'
        trans[0,0] = f'Matrix of transitions {name} (number of cases)'
        sheets = []
        for i, t in enumerate(thresholds):
            mkv, tr, t = markov(df, t, name)
            row = sum((lens+2)[:i])+2
            n = len(t)
            for x,m,s,perc in [(mrkov, mkv, ws_markov,  True), 
                                (trans, tr,  ws_trans,   False)]:
                x[row+1,0] = 'To interval'
                x[row-1,2] = 'From interval'
                x[row,n+2] = 'sum'
                x[row,2:n+2] = t
                x[row+1:row + n+1,1] = t 
                x[row+1:row + n+1,2:n+2] = m
                x[row+1:row + n+1,n+2] = np.sum(m,axis=1)
                populate_ws(s,x)
                sheets.append(s)
                set_fmt(s, 'font', Font(bold = True),row-1, 2, row+1, len(x[0]))
                if perc:
                    fmt = "0 %"
                else:
                    fmt = "_-* # ##0_-;-* # ##0_-;_-* \"-\"??_-;_-@_-"
                set_fmt(s, 'number_format', fmt,row+1, 2, row + n+1, n+2)

        for s in sheets:
            set_fmt(s, 'font', Font(bold = True), 0, 0, len(mrkov), 2)

    
    #populate_ws(wb.create_sheet('data'), df)
    wb.save(f'markov.xlsx')
                   
def populate_ws(ws, x):
    """Sets a range in the worksheet from a 2D list or numpy array."""
    if type(x)==pd.DataFrame:
        x = np.vstack((x.columns,np.array(x)))
    for i, row in enumerate(x):
        for j, value in enumerate(row):
            if value == '':
                value = None
            try:
                value = float(value)
            except:
                pass
            if value == float('inf'):
                value = 'inf'
            ws.cell(i+1, j+1).value = value
    return ws

def set_fmt(ws, property, fmt, row, col, endrow = None, endcol = None):
    if endcol is None:
        endcol = col +1
    if endrow is None:
        endrow = col +1
    for i in range(row, endrow):
        for j in range(col, endcol):
            setattr(ws.cell(i+1,j+1),property,fmt)

 
def markov(df, thresholds, varname):
    gs = []
    n = len(thresholds) +1
    #scaling:
    thresholds = np.array(thresholds)*np.max(df[varname])
    #includes boundaries
    thresholds = [0] + list(thresholds) + [np.inf]
    #creates a column of lower and a column of upper threashold for each group (rows)
    thresholds = np.column_stack((thresholds, np.roll(thresholds, -1)))[:-1]
    m = np.zeros((n,n))
    markov = np.zeros((n,n))
    for i, (t0,t1) in enumerate(thresholds):
        g = df[(df[varname]>=t0)&(df[varname]<t1)][varname + '_L1']
        gs.append(g)

        for j, (to_t0, to_t1) in enumerate(thresholds):
            m[i,j] = np.sum((g>=to_t0)&(g<to_t1))
            markov[i,j] = m[i,j]/len(g)

    return markov, m, thresholds[:,1]


    
def print_table(markov, stderr, heading):
    s = heading + "\n\t"
    s += '\t'*int(len(markov[0])*0.5) +'to \n'
    
    for i in range(len(markov)):
        if i==int(len(markov)*0.5):
            s += 'from'
        s += '\t' + '\t'.join(f"{markov[i,j]*100:.2f}%" for j in range(len(markov[i])))+ "\n"
        s += '\t' + '\t'.join(f"({stderr[i,j]*100:.2f}%)" for j in range(len(markov[i])))+ "\n"
    print(s)
    return s